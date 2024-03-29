#!/usr/bin/python3
# -*- coding: utf-8 -*-
"""
    pysic.io.load.py : function to import ice core data from ice core data spreadsheet
"""

import logging
import openpyxl
import pandas as pd
import numpy as np
import os

import pysic
from pysic.io import subvariable_dict, property2sheet
from pysic.io import update
from pysic.tools import inverse_dict, parse_datetimetz, parse_coordinate, isfloat
from pysic.core.core import __CoreVersion__
from pysic.io import find_str_in_col, find_str_in_row

MAX_ROW = 200
MAX_COL = 200
TOL = 1e-12

# DEGUG
logging.basicConfig(level=logging.DEBUG)
drop_empty = False
fill_missing = True
ic_property = None


# TODO: modifiy reading to account for false bottom

def list_folder(dirpath, fileext='.xlsx', level=0):
    """
    list all files with specific extension in a directory

    :param dirpath: str; directory to scan for ice core
    :param fileext: str, default .xlsx; file extension for ice core data
    :param level: numeric, default 0; level of recursitivy in directory search
    :return ic_list: list
        list of ice core path
    """

    if not fileext.startswith('.'):
        fileext = '.' + fileext

    _ics = []

    logger = logging.getLogger(__name__)

    def walklevel(some_dir, level=level):
        some_dir = some_dir.rstrip(os.path.sep)
        assert os.path.isdir(some_dir)
        num_sep = some_dir.count(os.path.sep)
        for root, dirs, files in os.walk(some_dir):
            yield root, dirs, files
            num_sep_this = root.count(os.path.sep)
            if num_sep + level <= num_sep_this:
                del dirs[:]

    for dirName, subdirList, fileList in walklevel(dirpath, level=level):
        _ics.extend([dirName + '/' + f for f in fileList if f.endswith(fileext)])

    ics_set = set(_ics)
    logger.info("Found %i ice core datafile in %s" % (ics_set.__len__(), dirpath))

    return ics_set

def ic_from_path(ic_path, ic_property=None, drop_empty=False, fill_missing=True):
    from pysic.core.profile import Profile
    """
    :param ic_path:
        string: path to the ice core data. Ice core data should be stored in defined spreadsheet
    :param ic_properties:
        list of string: properties to import. If not specified all properties are imported.
    :param drop_empty:
        boolean: True drop empty profile
    :param fill_missing:
        boolean, default True: If true, missing section are filled with np.nan/none
    :return:
    """
    from pysic.core.core import Core

    logger = logging.getLogger(__name__)

    wb = openpyxl.load_workbook(filename=ic_path, keep_vba=False)  # load the xlsx spreadsheet

    ws_summary = wb['metadata-station']
    ws_metadata_core = wb['metadata-core']

    name = ws_metadata_core['C1'].value
    if ws_summary['C1'].value:
        version = ws_summary['C1'].value
        if version != __CoreVersion__:
            logger.warning("%s: ice core data is updated from %s to %s" % (name, version, __CoreVersion__))
            # TODO: update function pysic.core.update_data_spreadsheet
            wb = update.ic_data(wb, 'N/A')
    else:
        logger.error("(%s) ice core spreadsheet version not unavailable" % name)
        wb.close()

    comments = []

    logger.info("Importing data for %s" % ic_path)
    # Import datetime with tz
    datetime = parse_datetimetz(ws_metadata_core['C2'].value, ws_metadata_core['C3'].value, ws_metadata_core['D3'].value)

    # project
    row_idx = find_str_in_col(ws_summary, 'expedition', 1)[0]
    expedition = ws_summary.cell(row_idx, 3).value
    site = ws_summary.cell(row_idx + 1, 3).value

    # Location
    row_idx = find_str_in_col(ws_summary, 'latitude', 1)[0]
    lat_start_deg = parse_coordinate(ws_summary.cell(row_idx, 3).value,
                                     ws_summary.cell(row_idx, 4).value,
                                     ws_summary.cell(row_idx, 5).value)
    lon_start_deg = parse_coordinate(ws_summary.cell(row_idx + 1, 3).value,
                                     ws_summary.cell(row_idx + 1, 4).value,
                                     ws_summary.cell(row_idx + 1, 5).value)
    lat_end_deg = parse_coordinate(ws_summary.cell(row_idx, 6).value,
                                     ws_summary.cell(row_idx, 7).value,
                                     ws_summary.cell(row_idx, 8).value)
    lon_end_deg = parse_coordinate(ws_summary.cell(row_idx + 1, 6).value,
                                     ws_summary.cell(row_idx + 1, 7).value,
                                     ws_summary.cell(row_idx + 1, 8).value)

    # Station time
    row_idx = find_str_in_col(ws_summary, 'date', 1)[0]
    station_date_start = parse_datetimetz(ws_summary.cell(row_idx, 3).value,
                                          ws_summary.cell(row_idx + 1, 3).value,
                                          ws_summary.cell(row_idx + 2, 3).value)
    station_date_end = parse_datetimetz(ws_summary.cell(row_idx, 4).value,
                                        ws_summary.cell(row_idx + 1, 4).value,
                                        ws_summary.cell(row_idx + 2, 4).value)

    # Snow Depth
    n_snow = 1
    row_idx = find_str_in_col(ws_summary, 'snow depth', 1)[0]
    snow_depth = []
    while ws_summary.cell(row=row_idx, column=3+n_snow).value is not None:
        snow_depth.append(ws_summary.cell(row=row_idx, column=3+n_snow).value)
        n_snow += 1
    snow_depth = np.array(snow_depth)
    if any(snow_depth) > 1:  # check for metric unit
        logger.warning('%s: check if snow depth are reported in cm rather than m' % name)

    # Snow depth measurement
    if len(snow_depth) > 0:
        if isinstance(snow_depth[-1], str):
            comments.append(snow_depth[-1])
            snow_depth = pd.to_numeric(snow_depth[:-1], errors='coerce')
        else:
            snow_depth = pd.to_numeric(snow_depth, errors='coerce')

    # Snow average
    if isinstance(ws_summary.cell(row=row_idx, column=3).value, (float, int)):
        snow_depth_avg = ws_summary.cell(row=row_idx, column=3).value
    elif not np.isnan(snow_depth).all():
        snow_depth_avg = np.nanmean(snow_depth)
    else:
        snow_depth_avg = [np.nan]
    if len(snow_depth) == 0:
        snow_depth = snow_depth_avg

    # Ice thickness
    h_i = read_metadata_variable_as_float(ws_metadata_core, 'ice thickness')
    if np.isnan(h_i):
        logger.info('%s ice thickness is not a number' % name)
    elif h_i > 10:  # check for metric unit
        logger.warning('%s: check if ice thickness is reported in cm rather than m' % name)

    # Ice draft
    h_d = read_metadata_variable_as_float(ws_metadata_core, 'draft')
    if np.isnan(h_d):
        logger.info('%s draft is not a number' % name)
    elif h_d > 10:  # check for metric unit
        logger.warning('%s: check if draft is reported in cm rather than m' % name)

    # Ice freeboard
    h_f = read_metadata_variable_as_float(ws_metadata_core, 'freeboard')
    if np.isnan(h_f):
        logger.info('%s freeboard is not a number' % name)
    elif h_f > 10:  # check for metric unit
        logger.warning('%s: check if freeboard is reported in cm rather than m' % name)

    if np.isnan(h_f) and (not (np.isnan(h_d) and np.isnan(h_i))):
        h_f = h_i - h_d
        logger.info('%s compute ice freeboard as ice thickness minus ice draft' % name)
    elif np.isnan(h_d) and (not (np.isnan(h_f) and np.isnan(h_i))):
        h_d = h_i - h_f
        logger.info('%s compute ice draft as ice thickness minus ice freeboard' % name)

    # Core length l_c (measured in with ruler)
    l_c = read_metadata_variable_as_float(ws_metadata_core, 'core length')
    if np.isnan(l_c):
        logger.info('%s core length is not a number' % name)
    elif l_c > 10:  # check for metric unit
        logger.warning('%s: check if core length is reported in cm rather than m' % name)

    core = Core(name, datetime, expedition, lat_start_deg, lon_start_deg, l_c, h_i, h_f, snow_depth)

    # Temperature values
    core.t_air = read_metadata_variable_as_float(ws_summary, 'air temperature')
    core.t_snow_surface = read_metadata_variable_as_float(ws_summary, 'snow surface temperature')
    core.t_ice_surface = read_metadata_variable_as_float(ws_summary, 'snow/ice temperature')
    core.t_water = read_metadata_variable_as_float(ws_summary, 'seawater temperature')
    core.s_water = read_metadata_variable_as_float(ws_summary, 'seawater salinity')

    # Sampling event
    core.station = read_metadata_variable_as_str(ws_summary, 'sampling station')
    core.protocol = read_metadata_variable_as_str(ws_summary, 'procedure')
    m_col = 3
    row_idx = find_str_in_col(ws_summary, 'associated cores (1 by cell)')[0]
    while ws_summary.cell(row_idx, m_col).value:
        core.add_to_collection(ws_summary.cell(row_idx, m_col).value)
        m_col += 1

    # Sampling instrument
    instrument_d = {}
    row_idx = find_str_in_col(ws_metadata_core, 'INSTRUMENTS')[0] + 1
    while ws_metadata_core.cell(row_idx, 1).value is not None:
        instrument_d[ws_metadata_core.cell(row_idx, 1).value] = ws_metadata_core.cell(row_idx, 3).value
        row_idx += 1
    core.instrument = instrument_d

    # Core collection
    m_col = 3
    row_collection = find_str_in_col(ws_summary, 'associated cores (1 by cell)')[0] + 1
    while ws_summary.cell(row_collection, m_col).value:
        core.add_to_collection(ws_summary.cell(row_collection, m_col).value)
        m_col += 1

    # comment
    comments.append(read_metadata_variable_as_str(ws_summary, 'GENERAL COMMENTS', row_offset=1))
    core.add_comment('; '.join(list(filter(lambda c: c not in ['N/A', None], comments))))

    # weather
    # TODO: read weather information

    # References
    reference_d = {}
    row_idx = find_str_in_col(ws_metadata_core, 'ICE')[0]
    if ws_metadata_core.cell(row_idx, 4).value is not None:
        if ws_metadata_core.cell(row_idx + 1, 4).value is not None:
            reference_d['ice'] = [ws_metadata_core.cell(row_idx, 4).value, ws_metadata_core.cell(row_idx + 1, 4).value]
        else:
            reference_d['ice'] = [ws_metadata_core.cell(row_idx, 4).value, 'down']
    else:
        reference_d['ice'] = [None, None]
        logger.info('Vertical reference for ice not defined')
    if ws_metadata_core.cell(row_idx + 2, 4).value is not None:
        if ws_metadata_core.cell(row_idx + 3, 4).value is not None:
            reference_d['snow'] = [ws_metadata_core.cell(row_idx + 2, 4).value, ws_metadata_core.cell(row_idx + 3, 4).value]
        else:
            reference_d['snow'] = [ws_metadata_core.cell(row_idx + 2, 4).value, 'up']
    else:
        reference_d['snow'] = [None, None]
        logger.info('Vertical reference for snow not defined')
    if ws_metadata_core.cell(row_idx + 4, 4).value is not None:
        if ws_metadata_core.cell(row_idx + 5, 4).value is not None:
            reference_d['seawater'] = [ws_metadata_core.cell(row_idx + 4, 4).value, ws_metadata_core.cell(row_idx + 5, 4).value]
        else:
            reference_d['seawater'] = [ws_metadata_core.cell(row_idx + 4, 4).value, 'down']
    else:
        reference_d['seawater'] = [None, None]
        logger.info('Vertical reference for seawater not defined')
    core.reference.update(reference_d)

    # PAR
    core.par_incoming = read_metadata_variable_as_float(ws_summary, 'PAR readings')
    par_unit = read_metadata_variable_as_str(ws_summary, 'PAR readings', col_variable_idx=2)
    core.unit.update({'par': par_unit})
    core.par_transmitted = read_metadata_variable_as_float(ws_summary, 'PAR readings', row_offset=1)
    par_unit = read_metadata_variable_as_str(ws_summary, 'PAR readings', col_variable_idx=2, row_offset=1)
    if par_unit != core.unit['par']:
        logger.error('PAR unit not consistent')

    # import property profile
    if ic_property is None:
        worksheets = [sheet for sheet in wb.sheetnames if (sheet not in ['tex', 'TM',  'summary', 'abbreviation', 'locations',
                                                              'lists', 'Vf_oil_calculation', 'metadata-core',
                                                              'metadata-station', 'density-volume', 'sediment',
                                                              'ct']) and
                  (sheet.lower().find('fig') == -1)]
        # Always import 'salo18' then 'temperature' before the other
        if 'temp' in worksheets:
            worksheets.remove('temp')
            worksheets = ['temp'] + worksheets
        if 'salo18' in worksheets:
            worksheets.remove('salo18')
            worksheets = ['salo18'] + worksheets

    else:
        worksheets = []
        if not isinstance(ic_property, list):
            if ic_property.lower().find('state variable')+1:
                ic_property = ['temperature', 'salinity']
            else:
                ic_property = [ic_property]

        _imported_variables = []
        for ic_prop in ic_property:
            if property2sheet[ic_prop] in wb.sheetnames and ic_prop not in _imported_variables:
                worksheets.append(property2sheet[ic_prop])

    for sheet in worksheets:
        ws_property = wb[sheet]
        if sheet == 'snow':
            profile = read_snow_profile(ws_property, ic_property=None, reference_d=reference_d)
            matter = 'snow'
        elif sheet == 'seawater':
            profile = read_generic_profile(ws_property, ic_property=None, reference_d=reference_d, core_length=core.length, fill_missing=fill_missing)
            matter = 'seawater'
        elif sheet == 'sackhole' or sheet == 'brine':
            profile = read_generic_profile(ws_property, ic_property=None, reference_d=reference_d, core_length=core.length, fill_missing=fill_missing)
            matter = 'brine'
        else:
            matter = 'ice'
            profile = read_generic_profile(ws_property, ic_property=None, reference_d=reference_d, core_length=core.length, fill_missing=fill_missing)

        if not profile.empty:
            profile['matter'] = matter
            # TODO def add_sw_salinity(profile):
            # add sea water salinity at the correct
            if matter == 'ice' and 'salinity' in profile.variable and not np.isnan(core.s_water):
                # TODO: def profile.get_vref():
                v_ref_loc = profile.v_ref_loc.unique()
                if len(v_ref_loc) != 1:
                    logger.error('%s\t\tvertical reference location not unique')
                else:
                    v_ref_loc = v_ref_loc[0]
                    v_ref_dir = profile.v_ref_dir.unique()
                    if len(v_ref_dir) != 1:
                        logger.error('%s\t\tvertical reference direction not unique')
                    else:
                        v_ref_dir = v_ref_dir[0]
                        v_ref_h = profile.v_ref_h.unique()
                        if len(v_ref_h) != 1:
                            logger.error('%s\t\tvertical reference height not unique')
                        else:
                            v_ref_h = v_ref_h[0]
            #    headers = ['y_low', 'y_sup', 'salinity_ID', 'salinity_value', 'salinity_quality', 'v_ref_loc', 'v_ref_dir', 'v_ref_h', 'matter']
            # if v_ref_loc == 'ice bottom':
            #     # y_low =
            #     # y_sup =
            #     data = [y_low, y_sup, 's_sw', core.s_water, 0, v_ref_loc, v_ref_dir, v_ref_h, 'seawater']
            # elif v_ref_loc == 'ice surface':
            #     # y_low =
            #     # y_sup =
            #     data = [y_low, y_sup, 's_sw', core.s_water, 0, v_ref_loc, v_ref_dir, v_ref_h, 'seawater']
            # else:
            #     logger.error('%s\t\tTODO: implement v_ref_loc %s for profile%' %(core.name, v_ref_loc))

            # add snow, ice surface, and sea water temperature in temperature profile
            if matter == 'ice' and 'temperature' in profile.variable:
                headers = ['y_mid', 'temperature_value', 'temperature_quality', 'comment', 'matter', 'property', 'v_ref_loc',
                           'v_ref_h', 'v_ref_dir']

                v_ref_loc = profile.v_ref_loc.unique()
                v_ref_h = profile.v_ref_h.unique()
                v_ref_dir = profile.v_ref_dir.unique()

                if len(v_ref_loc) == 1 and len(v_ref_h) == 1 and len(v_ref_dir) == 1:
                    v_ref_loc = v_ref_loc[0]
                    v_ref_h = v_ref_h[0]
                    v_ref_dir = v_ref_dir[0]
                    if v_ref_dir == 'positive':
                        coef_dir = 1
                    else:
                        coef_dir = -1

                    if v_ref_h != 0:
                        logger.error('%s - %s: v_ref_h is not 0 ' % (core.name, sheet))
                    elif v_ref_loc == 'ice surface':
                        if not -1 in profile.y_mid.values:
                            # air temperature 1 m above snow surface
                            if isinstance(core.t_air, (float, int)) and not np.isnan(core.t_air):
                                data_surface = [coef_dir*-1, core.t_air, 1, 'Air temperature', 'air', 'temperature', v_ref_loc, v_ref_h, v_ref_dir]
                                profile = profile.append(pd.DataFrame([data_surface], columns=headers))
                        if not 0 in profile.y_mid.values:
                            # ice surface
                            if isinstance(core.t_ice_surface, (float, int)) and not np.isnan(core.t_ice_surface):
                                data_surface = [0, core.t_ice_surface, 1, 'Ice surface temperature', 'ice', 'temperature', v_ref_loc, v_ref_h, v_ref_dir]
                                profile = profile.append(pd.DataFrame([data_surface], columns=headers))
                            # snow surface
                            if isinstance(core.t_snow_surface, (float, int)) and not np.isnan(core.t_snow_surface) and isinstance(core.snow_depth, (float, int)) and not np.isnan(core.snow_depth):
                                data_snow = [coef_dir*core.snow_depth, core.t_air, 1, 'Snow surface temperature', 'snow', 'temperature', v_ref_loc, v_ref_h, v_ref_dir]
                                profile = profile.append(pd.DataFrame([data_snow], columns=headers))
                        if core.length - profile.y_mid.max() > TOL:
                            # ice bottom / seawater
                            if isinstance(core.t_water, (float, int)) and not np.isnan(core.t_water):
                                data_bottom = [coef_dir*core.length, core.t_water, 1, 'Seawater temperature', 'ice', 'temperature', v_ref_loc, v_ref_h, v_ref_dir]
                                profile = profile.append(pd.DataFrame([data_bottom], columns=headers))
                    elif v_ref_loc == 'ice bottom':
                        if not 0 in profile.y_mid.values:
                            # ice bottom
                            if isinstance(core.t_water, (float, int)) and not np.isnan(core.t_water):
                                data_bottom = [0, core.t_water, 1, 'Seawater temperature', 'ice', 'temperature', v_ref_loc, v_ref_h, v_ref_dir]
                                profile = profile.append(pd.DataFrame([data_bottom], columns=headers))
                        # TODO: which one to choose core.length or profile.y_mid.max():
                        # select max(y_mid) if max(y_mid) > core.lenght
                        if np.abs(core.length - profile.y_mid.max()) < TOL:
                            # ice surface
                            if isinstance(core.t_ice_surface, (float, int)) and not np.isnan(core.t_ice_surface):
                                data_surface = [coef_dir*core.length, core.t_ice_surface, 1, 'Ice surface temperature', 'ice', 'temperature', v_ref_loc, v_ref_h, v_ref_dir]
                                profile = profile.append(pd.DataFrame([data_surface], columns=headers))
                            # snow surface
                            if isinstance(core.t_snow_surface, (float, int)) and not np.isnan(core.t_snow_surface) and isinstance(core.snow_depth, (float, int)) and not np.isnan(core.snow_depth):
                                data_snow = [coef_dir*core.snow_depth+coef_dir*core.length, core.t_air, 1, 'Snow surface temperature', 'snow', 'temperature', v_ref_loc, v_ref_h, v_ref_dir]
                                profile = profile.append(pd.DataFrame([data_snow], columns=headers))
                            # air temperature
                            if isinstance(core.t_air, (float, int)) and not np.isnan(core.t_air):
                                data_surface = [coef_dir*1+coef_dir*core.length, core.t_air, 1, 'Air temperature', 'air', 'temperature', v_ref_loc, v_ref_h, v_ref_dir]
                                profile = profile.append(pd.DataFrame([data_surface], columns=headers))
                        else:
                            # ice surface
                            if isinstance(core.t_ice_surface, (float, int)) and not np.isnan(core.t_ice_surface):
                                data_surface = [coef_dir*profile.y_mid.max(), core.t_ice_surface, 1, 'Ice surface temperature', 'ice', 'temperature', v_ref_loc, v_ref_h, v_ref_dir]
                                profile = profile.append(pd.DataFrame([data_surface], columns=headers))
                            # snow surface
                            if isinstance(core.t_snow_surface, (float, int)) and not np.isnan(core.t_snow_surface) and isinstance(core.snow_depth, (float, int)) and not np.isnan(core.snow_depth):
                                data_snow = [coef_dir * core.snow_depth + coef_dir * profile.y_mid.max(), core.t_air, 1,
                                             'Snow surface temperature', 'snow', 'temperature', v_ref_loc,
                                             v_ref_h, v_ref_dir]
                                profile = profile.append(pd.DataFrame([data_snow], columns=headers))
                            # air temperature
                            if isinstance(core.t_air, (float, int)) and not np.isnan(core.t_air):
                                data_surface = [coef_dir*1+coef_dir*profile.y_mid.max(), core.t_air, 1, 'Air temperature', 'air', 'temperature', v_ref_loc, v_ref_h, v_ref_dir]
                                profile = profile.append(pd.DataFrame([data_surface], columns=headers))
                else:
                    logger.error('%s - %s: vertical references mixed up ' %(core.name, sheet))
                profile = profile.sort_values(by='y_mid')
                profile = profile.reset_index(drop=True)
            profile['name'] = name
            if drop_empty:
                profile.drop_empty_property()

            profile = pysic.Profile(profile)
            if not profile.empty:
                core.add_profile(profile)
                logger.info('(%s) data imported with success: %s' % (core.name, ", ".join(profile.variable)))
            else:
                logger.info('(%s) no data to import from %s ' % (core.name, sheet))
    # else:
    #     if not isinstance(ic_property, list):
    #         if ic_property.lower().find('state variable')+1:
    #             ic_property = ['temperature', 'salinity']
    #         else:
    #             ic_property = [ic_property]
    #
    #     _imported_variables = []
    #     for ic_prop in ic_property:
    #         if property2sheet[ic_prop] in sheets and ic_prop not in _imported_variables:
    #             sheet = property2sheet[ic_prop]
    #             ws_property = wb[sheet]
    #             property2import = [p for p in ic_property if p in inverse_dict(property2sheet)[sheet]]
    #             if sheet == 'snow':
    #                 profile = read_generic_profile(ws_property, ic_property=None, reference_d=reference_d)
    #                 matter = 'snow'
    #             elif sheet == 'seawater':
    #                 matter = 'seawater'
    #                 profile = read_generic_profile(ws_property, ic_property=None, reference_d=reference_d, fill_missing=True)
    #             elif sheet == 'sackhole' or sheet == 'brine':
    #                 matter = 'brine'
    #                 profile = read_generic_profile(ws_property, ic_property=None, reference_d=reference_d, fill_missing=True)
    #             else:
    #                 matter = 'ice'
    #                 profile = read_generic_profile(ws_property, ic_property=None, reference_d=reference_d, fill_missing=True)
    #
    #             if not profile.empty:
    #                 profile['matter'] = matter
    #         core.profile = Profile(pd.concat([core.profile, profile]))

    # Add air, snow surface, ice surface, seawater temperature to temperature profile if exist


    # Add seawater salinity to salinity profile

        #
        #                 # Add temperature at ice surface for temperautre profile
        #                 if matter == 'ice' and profile.get_property() is not None and 'temperature' in profile.get_property():
        #                     headers = ['y_mid', 'temperature_value', 'comment', 'variable', 'v_ref', 'matter']
        #                     v_ref = profile.v_ref.unique()
        #                     if len(v_ref) == 1:
        #                         v_ref = v_ref[0]
        #                         if v_ref == 'top':
        #                             if not -1 in profile.y_mid.values:
        #                                 # air temperature 1 m above snow surface
        #                                 if isinstance(core.t_air, (float, int)) and not np.isnan(core.t_air):
        #                                     data_surface = [-1, core.t_air, 'Air temperature', 'temperature', 'top', 'air']
        #                                     profile = profile.append(pd.DataFrame([data_surface], columns=headers))
        #                             if not 0 in profile.y_mid.values:
        #                                 # ice surface
        #                                 if isinstance(core.t_ice_surface, (float, int)) and not np.isnan(core.t_ice_surface):
        #                                     data_surface = [0, core.t_ice_surface, 'Ice surface temperature', 'temperature',
        #                                                     'top', 'ice']
        #                                     profile = profile.append(pd.DataFrame([data_surface], columns=headers))
        #                             if core.length - profile.y_mid.max() > TOL:
        #                                 # ice bottom / seawtaer
        #                                 if isinstance(core.t_water, (float, int)) and not np.isnan(core.t_water):
        #                                     data_bottom = [core.length, core.t_water, 'Ice bottom temperature', 'temperature',
        #                                                    'top', 'ice']
        #                                     profile = profile.append(pd.DataFrame([data_bottom], columns=headers))
        #                                     data_bottom = [core.length, core.t_water, 'Seawater temperature', 'temperature',
        #                                                    'top', 'seawater']
        #                                     profile = profile.append(pd.DataFrame([data_bottom], columns=headers))
        #                         elif v_ref == 'bottom':
        #                             # air temperature
        #                             if isinstance(core.t_air, (float, int)) and not np.isnan(core.t_air):
        #                                 data_surface = [1, core.t_air, 'Air temperature', 'temperature', 'bottom', 'air']
        #                                 profile = profile.append(pd.DataFrame([data_surface], columns=headers))
        #                             if not 0 in profile.y_mid.values:
        #                                 # ice bottom
        #                                 if isinstance(core.t_water, (float, int)) and not np.isnan(core.t_water):
        #                                     data_bottom = [0, core.t_water, 'Seawater temperature', 'temperature', 'bottom',
        #                                                    'ice']
        #                                     profile = profile.append(pd.DataFrame([data_bottom], columns=headers))
        #                             if np.abs(core.length - profile.y_mid.max()) < TOL:
        #                                 # ice surface
        #                                 if isinstance(core.t_ice_surface, (float, int)) and not np.isnan(core.t_ice_surface):
        #                                     data_surface = [core.length, core.t_ice_surface, 'Ice surface temperature',
        #                                                     'temperature', 'bottom', 'ice']
        #                                     profile = profile.append(pd.DataFrame([data_surface], columns=headers))
        #                     else:
        #                         logger.error('%s - %s: vertical references mixed up ' % (core.name, sheet))
        #                     profile = profile.sort_values(by='y_mid')

    return core


def read_generic_profile(ws_property, ic_property=None, reference_d={'ice': ['ice surface', 'down']}, core_length=np.nan, fill_missing=True, drop_empty=False):
    """
    :param ws_property:
        openpyxl.worksheet, worksheet property to import
    :param ic_property:
        str array, containing property to import. Default None import all properties available
    :param reference_d:
        dict, contain information of vertical reference system
    :param core_length:
        float, default np.nan (optional). Ice core length
    :param fill_missing:
    :param drop_empty:
    """
    from pysic.core.profile import Profile
    from pysic.property import prop_associated

    logger = logging.getLogger(__name__)

    # find last column number with column header and/or subheaders
    if ws_property.max_row < MAX_ROW:
        max_row = ws_property.max_row
    else:
        max_row = MAX_ROW
    if ws_property.max_column < MAX_COL:
        max_col = ws_property.max_column
    else:
        max_col = MAX_COL
    min_row = 4

    # Dictionnary
    # parse property headers:
    header_d = {}
    header_unit_d = {}

    n_row = 1  # header row
    n_col_min = 1  # start column
    n_col = n_col_min
    empty_header = 0
    while empty_header < 5 and n_col < max_col:
        # Read depth
        if isinstance(ws_property.cell(n_row, n_col).value, str):
            if 'depth' in ws_property.cell(n_row, n_col).value:
                h_ = ws_property.cell(n_row, n_col).value
                hs_ = ws_property.cell(n_row + 1, n_col).value
                hu_ = ws_property.cell(n_row + 2, n_col).value
                if h_ not in header_d:
                    header_d[h_] = {hs_: n_col}
                    header_unit_d[h_] = {hs_: hu_}
                else:
                    header_d[h_].update({hs_: n_col})
                    header_unit_d[h_].update({hs_: hu_})
            elif ws_property.cell(n_row, n_col).value == 'comment':
                h_ = ws_property.cell(n_row, n_col).value
                hs_ = ws_property.cell(n_row + 1, n_col).value
                hu_ = ws_property.cell(n_row + 2, n_col).value
                header_d[h_] = {hs_: n_col}
                header_unit_d[h_] = {hs_: hu_}

            # specific reader for temperature
            elif ws_property.cell(n_row, n_col).value == 'temperature':
                h_ = ws_property.cell(n_row, n_col).value
                hs_ = ws_property.cell(n_row + 1, n_col).value
                hu_ = ws_property.cell(n_row + 2, n_col).value
                if h_ not in header_d:
                    header_d[h_] = {hs_: n_col}
                    header_unit_d[h_] = {hs_: hu_}
                else:
                    header_d[h_].update({hs_: n_col})
                    header_unit_d[h_] = {hs_: hu_}

                hs_ = ws_property.cell(n_row + 1, n_col+1).value
                hu_ = ws_property.cell(n_row + 2, n_col+1).value
                header_d[h_].update({hs_: n_col+1})
                header_unit_d[h_] = {hs_: hu_}

            # specific reader entry for eco_pool tab
            else:
                prop_col = n_col
                new_prop = False
                qual_col = None
                while not new_prop and empty_header < 5:
                    if isinstance(ws_property.cell(1, prop_col).value, str):
                        if 'ID' in header_d[h_] and 'quality' not in header_d[h_]:
                            if qual_col is not None:
                                header_d[h_].update({'quality': qual_col})
                            else:
                                _col = n_col
                                while ws_property.cell(n_row + 1,
                                                       _col).value != 'quality' and _col < ws_property.max_column:
                                    _col += 1
                                qual_col = _col
                                if qual_col is not None:
                                    header_d[h_].update({'quality': qual_col})
                                else:
                                    logger.error(
                                        'pysic.load.read_generic_profile: undefined quality column for property %s' % h_)
                        h_ = ws_property.cell(n_row, prop_col).value
                        hs_ = ws_property.cell(n_row + 1, prop_col).value
                        hu_ = ws_property.cell(n_row + 2, prop_col).value
                        header_d[h_] = {hs_: prop_col}
                        header_unit_d[h_] = {hs_: hu_}
                    elif ws_property.cell(2, prop_col).value == 'quality':
                        qual_col = prop_col
                        hs_ = ws_property.cell(n_row + 1, qual_col).value
                        header_d[h_].update({hs_: prop_col})
                    elif ws_property.cell(n_row + 1, prop_col).value != 'ID':
                        if ws_property.cell(n_row + 1, prop_col).value is not None:
                            hs_ = ws_property.cell(n_row + 1, prop_col).value
                            header_d[h_].update({hs_: prop_col})
                        else:
                            empty_header += 1
                    prop_col += 1

                    _ID_flag = (ws_property.cell(n_row + 1, prop_col).value == 'ID')
                    _comment_flag = (ws_property.cell(n_row, prop_col).value == 'comment')
                    _ish_flag =  (ws_property.cell(n_row, prop_col).value == 'ice section height')
                    if _ID_flag or _comment_flag or  _ish_flag:
                        if 'ID' in header_d[h_] and 'quality' not in header_d[h_]:
                            header_d[h_].update({'quality': qual_col})
                        new_prop = True
                        n_col = n_col + (prop_col - n_col -1)
            n_col += 1
        # search for subsamples ID
        else:
            prop_col = n_col
            new_prop = False
            h_ = None
            subheader_d = {}
            subheader_unit_d = {}
            while not new_prop and empty_header < 5 and prop_col < max_col:
                if isinstance(ws_property.cell(1, prop_col).value, str):
                    h_ = ws_property.cell(n_row, prop_col).value
                    hs_ = ws_property.cell(n_row + 1, prop_col).value
                    hu_ = ws_property.cell(n_row + 2, prop_col).value
                    header_d[h_] = {hs_: prop_col}
                    header_unit_d[h_] = {hs_: hu_}

                    if 'quality' not in header_d[h_] or 'quality' not in subheader_d:
                        _col = prop_col
                        while not ws_property.cell(n_row + 1, _col).value == 'ID' and _col < ws_property.max_column:
                            if ws_property.cell(n_row + 1, _col).value == 'quality':
                                subheader_d['quality'] = _col
                                subheader_unit_d['quality'] = ws_property.cell(n_row + 2, _col).value
                                break
                            _col += 1

                    for key in subheader_d:
                        header_d[h_].update({key: subheader_d[key]})
                        header_unit_d[h_].update({key: subheader_unit_d[key]})

                elif h_ is None:
                    hs_ = ws_property.cell(n_row + 1, prop_col).value
                    hu_ = ws_property.cell(n_row + 2, prop_col).value
                    subheader_d[hs_] = prop_col
                    subheader_unit_d[hs_] = hu_
                else:
                    hs_ = ws_property.cell(n_row + 1, prop_col).value
                    hu_ = ws_property.cell(n_row + 2, prop_col).value
                    header_d[h_].update({hs_: prop_col})
                    header_unit_d[h_].update({hs_: hu_})
                    for key in subheader_d:
                        header_d[h_].update({key: subheader_d[key]})
                        header_unit_d[h_].update({key: subheader_unit_d[key]})
                prop_col += 1

                if ws_property.cell(n_row + 1, prop_col).value == 'ID' or ws_property.cell(n_row, prop_col).value == 'comment':
                    new_prop = True
            n_col = prop_col
            del subheader_d, subheader_unit_d, prop_col, new_prop
    del empty_header, n_col, n_row, hs_, hu_, h_

    # Read center depth for continuous profile and step profile (if available)
    if 'depth center' in header_d:
        logger.error('header not define')

    if 'depth' in header_d:
        loc3 = header_d['depth']['value']
        header_d['y_mid'] = header_d.pop('depth')

        # look for 1st numerical value
        for ii_row in range(min_row, max_row):
            min_row_3 = ii_row
            if isinstance(ws_property.cell(ii_row, loc3 + 1).value, (float, int)):
                break

        y_mid = np.array(
            [ws_property.cell(row, loc3).value for row in range(min_row_3, max_row)]).astype(float)

        # discard trailing nan value starting at the end
        y_nan_loc = [ii for ii in np.arange(1, len(y_mid))[::-1] if np.isnan(y_mid[ii])]
        y_mid = np.delete(y_mid, y_nan_loc)

        # if y_mid is not a numeric, then there is no data
        if all(np.isnan(y_mid)):
            y_mid = []
            min_row_3 = np.nan
    else:
        y_mid = []
        min_row_3 = np.nan

    # Read lower and upper section depth for step profile or set to nan for continuous profile
    if 'depth 1' in header_d and 'depth 2' in header_d:
        y_low_col = header_d['depth 1']['value']
        header_d['y_low'] = header_d.pop('depth 1')
        y_sup_col = header_d['depth 2']['value']
        header_d['y_sup'] = header_d.pop('depth 2')

        # find first numerical value
        for ii_row in range(min_row, max_row):
            if isinstance(ws_property.cell(ii_row, y_low_col).value, (float, int)):
                min_row_1 = ii_row
                break
            else:
                min_row_1 = np.nan
        for ii_row in range(min_row, max_row):
            if isinstance(ws_property.cell(ii_row, y_sup_col).value, (float, int)):
                min_row_2 = ii_row
                break
            else:
                min_row_2 = np.nan

        min_row_12 = min(min_row_1, min_row_2)
        if np.isnan(min_row_12):
            return Profile()
        y_low = np.array([ws_property.cell(ii_row, y_low_col).value for ii_row in range(min_row_12, max_row)]).astype(float)
        y_sup = np.array([ws_property.cell(ii_row, y_sup_col).value for ii_row in range(min_row_12, max_row)]).astype(float)

        # discard trailing nan value starting at the end
        y_nan_loc = [ii for ii in np.arange(1, len(y_low))[::-1] if np.isnan(y_low[ii]) and np.isnan(y_sup[ii])]
        y_low = np.delete(y_low, y_nan_loc)
        y_sup = np.delete(y_sup, y_nan_loc)

        if len(y_mid) == len(y_low) and (min_row_3 == min_row_12 or np.isnan(min_row_3)):
            if np.any(np.abs(((y_low + y_sup) / 2) - y_mid > TOL)):
                logger.error('\t\t%s: y_mid are not mid points of y_low and y_sup. Computing y_mid = (y_low+y_sup)/2'
                             % ws_property.title)
                y_mid = (y_low + y_sup) / 2
            else:
                logger.info('\t\t%s: y_mid are mid points of y_low and y_sup. Do nothing' % ws_property.title)
        elif np.isnan(y_mid).any():
            y_mid = (y_low + y_sup) / 2
            logger.info('\t\t%s: y_mid do not exit. Computing y_mid = (y_low+y_sup)/2'
                        % ws_property.title)
        elif min_row_3 != min_row_12:
            y_mid = (y_low + y_sup) / 2
            logger.info('\t\t%s: not all y_mid exit. Computing y_mid = (y_low+y_sup)/2'
                           % ws_property.title)
        else:
            y_mid = (y_low + y_sup) / 2
            logger.info('\t\t%s: y_mid does not exist. Computing y_mid = (y_low+y_sup)/2'
                        % ws_property.title)
    elif 'depth 1' in header_d:
        y_low_col = header_d['depth 1']['value']
        header_d['y_low'] = header_d.pop('depth 1')

        # find first numerical value
        for ii_row in range(min_row, max_row):
            if isinstance(ws_property.cell(ii_row, y_low_col).value, (float, int)):
                min_row_1 = ii_row
                break
        y_low = np.array([ws_property.cell(ii_row, y_low_col).value for ii_row in range(min_row_1, max_row + 1)]).astype(float)

        # discard trailing nan value starting at the end
        y_nan_loc = [ii for ii in np.arange(1, len(y_low))[::-1] if np.isnan(y_low[ii])]
        y_low = np.delete(y_low, y_nan_loc)

        # Fill y_sup
        logger.warning('\t\t%s: y_sup do not exist. Attempting to infer y_sup from y_low'
                       % ws_property.title)
        if not np.isnan(core_length):
            logger.info('\t\t%s: using core length, y_sup[-1] = l_c' % ws_property.title)
            y_sup = np.concatenate([y_low[1:], [core_length]])
        elif min_row_1 == min_row_3 and len(y_low) == len(y_mid):
            logger.warning('\t\t%s: using y_mid and y_low,  y_sup[-1] = y_low[-1] + 2 * (y_mid[-1] - y_low[-1])'
                           % ws_property.title)
            dy = 2 * (y_mid[-1] - y_low[-1])
            y_sup = np.concatenate([y_low[1:], [y_low[-1]+dy]])
        else:
            logger.warning('\t\t%s: core length not available, y_sup[-1] = y_low[-1] + (y_low[-1]-y_low[-2])'
                           % ws_property.title)
            dy = np.diff(y_low[-2:])
            y_sup = np.concatenate([y_low[1:], [y_low[-1]+dy]])
        min_row_12 = min_row_1
    elif 'depth 2' in header_d:
        y_sup_col = header_d['depth 2']['value']
        header_d['y_sup'] = header_d.pop('depth 2')

        # find first numerical value
        for ii_row in range(min_row, max_row):
            if isinstance(ws_property.cell(ii_row, y_sup_col).value, (float, int)):
                min_row_2 = ii_row
                break
        y_sup = np.array([ws_property.cell(ii_row, y_sup_col).value for ii_row in range(min_row_2, max_row + 1)]).astype(float)

        logger.warning('\t\t%s: y_low do not exist. Attempting to infer y_low from y_sup'% ws_property.title)
        # discard trailing nan value starting at the end
        y_nan_loc = [ii for ii in np.arange(1, len(y_sup))[::-1] if np.isnan(y_sup[ii])]
        y_sup = np.delete(y_sup, y_nan_loc)

        # Fill y_low
        if min_row_2 == min_row_3 and len(y_sup) == len(y_mid):
            dy = 2 * (y_sup[0] - y_mid[0])
            y_sup0 = y_sup[0] - dy
            if y_sup0 >= 0:
                logger.warning('\t\t%s: using y_mid and y_low,  y_low[0] = y_sup[0] - 2 * (y_sup[0]-y_mid[0])'
                               % ws_property.title)
                y_low = np.concatenate([[y_sup0], y_sup[1:]])
            else:
                logger.info('\t\t%s: For lower y_low, using y_low[0] = 0' % ws_property.title)
                y_low = np.concatenate([[0], y_sup[:-1]])
        else:
            logger.info('\t\t%s: For lower y_low, using y_low[0] = 0' % ws_property.title)
            y_low = np.concatenate([[0], y_sup[:-1]])
        min_row_12 = min_row_2
    else:
        logger.info('\t\t%s: y_low and y_sup do not exist. Creating nan array of same size as y_mid'% (ws_property.title))
        y_low = np.nan * np.ones(y_mid.__len__())
        y_sup = np.nan * np.ones(y_mid.__len__())
        min_row_12 = np.nan

    # Check length consistency
    if len(y_low) != len(y_mid):
        logger.error('\t\t%s: y_low/y_sup and y_mid of different size'% (ws_property.title))

    # Read data, according to depth value
    # look up for first numeric or standard entry value
    if not np.isnan(min_row_12) and not np.isnan(min_row_3):
        row_min = min(min_row_12, min_row_3)
    elif not np.isnan(min_row_12):
        row_min = min_row_12
    elif not np.isnan(min_row_3):
        row_min = min_row_3
    else:
        logger.info('\t\t%s: no data' % (ws_property.title))
        return Profile()

    row_max = row_min + len(y_mid) - 1
    n_col_max = n_col_min
    for key in header_d:
        for subkey in header_d[key]:
            if n_col_max < header_d[key][subkey]:
                n_col_max = header_d[key][subkey]

    # Read data by column
    _data = [[cell.value for cell in col] for col in ws_property.iter_cols(n_col_min, n_col_max, row_min, row_max)]

    # define matter type
    if ic_property is None:
        variable_prefix = ''
        matter = 'seaice'
    elif any(map(ic_property.__contains__, ['brine', 'sackhole'])):
        variable_prefix = 'brine_'
        matter = 'brine'
    elif any(map(ic_property.__contains__, ['seawater'])):
        variable_prefix = 'seawater_'
        matter = 'seawater'
    elif any(map(ic_property.__contains__, ['snow'])):
        variable_prefix = 'snow_'
        matter = 'snow'
    else:
        variable_prefix = ''
        matter = 'seaice'

    profile = pd.DataFrame()

#    for header in dict(sorted(header_d.items(), key=lambda item: item[1]['value'])):
    for header in dict(sorted(header_d.items(), key=lambda item: min(item[1].values()))):
        for subheader in header_d[header]:
            if len(header_d[header]) > 1:
                header_name = header + '_' + subheader
            else:
                header_name = header
            if profile.empty:
                profile = pd.DataFrame(_data[header_d[header][subheader]-1], columns=[header_name])
            else:
                profile = pd.concat([profile, pd.DataFrame(_data[header_d[header][subheader]-1], columns=[header_name])], axis=1)

    # Add 'y_mid' column if does not exist, and fill it
    if 'y_low' not in profile.keys():
        profile['y_low'] = y_low
    if 'y_mid' not in profile.keys():
        profile['y_mid'] = y_mid
    if 'y_sup' not in profile.keys():
        profile['y_sup'] = y_sup

    # Add 'comment' column if does not exist and fill it with none value
    if 'comment' not in profile.columns:
        profile['comment'] = None

    # # Fill row with np.nan/None for missing section
    # if fill_missing:
    #     for ii_row in np.where(profile.y_sup.values[:-1] - profile.y_low.values[1:] < -TOL)[0]:
    #         y_low = profile.loc[profile.index == ii_row, 'y_sup'].values
    #         y_sup = profile.loc[profile.index == ii_row + 1, 'y_low'].values
    #         ID_columns = [c for c in profile.columns if '_ID' in c]
    #         na_columns = [c for c in profile.columns if 'ID' not in c]
    #         empty_row = pd.DataFrame([[np.nan] * len(na_columns) + ['nan_ID'] * len(ID_columns)], columns=na_columns + ID_columns)
    #         empty_row['y_low'] = y_low
    #         empty_row['y_sup'] = y_sup
    #         empty_row['y_mid'] = (y_low + y_sup) / 2
    #         profile = pd.concat([profile, empty_row]).reset_index(drop=True)
    #         logger.info('\t\t%s: filling missing section (%.3f - %.3f) with nan/none values'
    #                     % (ws_property.title, y_low, y_sup))
    #     if any(np.isnan(profile.y_mid)):
    #         profile = profile.drop('y_mid', axis=1)
    #         profile = profile.sort_values('y_low').reset_index(drop=True)
    #     else:
    #         profile = profile.sort_values('y_mid').reset_index(drop=True)

    # Drop Empty headers
    if None in profile.columns:
        profile = profile.drop(labels=[None], axis=1)

    # get profile property from headers (e.g. salinity, temperature, ...)
    ic_property = [h.split('_ID')[0] for h in profile.columns if 'ID' in h]
    ic_property += [h.split('_value')[0] for h in profile.columns if '_value' in h]
    ic_property = list(set(ic_property))

    #
    type_string_header = ['comment']
    type_float_header = [h for h in profile.columns if h not in type_string_header and 'ID' not in h]
    type_string_header = [h for h in profile.columns if h not in type_float_header]

    # generate list of headers, with float type, and set column type to float
    profile[type_float_header] = profile[type_float_header].apply(pd.to_numeric, errors='coerce')

    # Drop property profile without ID and values
    for header in header_d:
        if 'value' in header_d[header] and 'ID' in header_d[header]:
            if all(profile[header + '_ID'].isna()) and all(profile[header + '_value'].isna()):
                logger.info('\t\t%s: dropping %s profile without ID or values' % (ws_property.title, header))
                for subheader in header_d[header]:
                    profile = profile.drop(labels=[header + '_' + subheader], axis=1)
                    if header in ic_property:
                        ic_property.remove(header)

    # # remove empty line if all element of depth are nan:
    # if not fill_missing:
    #     subset = [col for col in ['y_low', 'y_sup', 'y_mid'] if col in profile.columns]
    #     profile = profile.dropna(axis=0, subset=subset, how='all')

    # set property by row
    def add_property(x, ic_prop):
        if x is None:
            x = ic_prop
        else:
            x = ', '.join(list(set(filter(None, x.split(', ')))) + [ic_prop])
        return x
    profile['property'] = None

    # TODO: add property to profile even if na (override)
    for ic_prop in ic_property:
        # For property with associated property e.g. salinity with conductivty and specific conductance, consider only
        try:
            profile_index = profile[(profile[ic_prop + '_ID'].notna() | profile[ic_prop + '_value'].notna())].index
        except KeyError:
            try:
                profile_index = profile[(profile[ic_prop + '_ID'].notna())].index
            except KeyError:
                profile_index = profile[(profile[ic_prop + '_value'].notna())].index
        else:
            pass
        profile.loc[profile_index, 'property'] = profile.loc[profile_index, 'property'].apply(lambda x: add_property(x, ic_prop))

    # set location, direction and depth of the vertical referential system
    # TODO: improve vertical reference
    if reference_d['ice'][0] == 'ice surface':
        v_ref_loc = 'ice surface'
        v_ref_h = 0
    elif reference_d['ice'][0] == 'ice/water interface':
        v_ref_loc = 'ice bottom'
        v_ref_h = 0
    else:
        logger.error('%s\t\tVertical reference %s not defined.' % (ws_property, reference_d['ice'][0]))

    profile['v_ref_loc'] = [v_ref_loc] * len(profile.index)
    profile['v_ref_h'] = [v_ref_h] * len(profile.index)
    type_string_header.append(v_ref_loc)
    type_float_header.append(v_ref_h)

    if reference_d['ice'][1] == 'up':
        v_ref_dir = 'positive'
    elif reference_d['ice'][1] == 'down':
        v_ref_dir = 'negative'
    else:
        logger.error('%s\t\tVertical reference %s not defined.' % (ws_property, reference_d['ice'][0]))

    profile['v_ref_dir'] = [v_ref_dir] * len(profile.index)
    type_float_header.append(v_ref_dir)

#    profile[type_string_header] = profile[type_string_header].astype(str).replace({'nan': None})
    return Profile(profile)


def read_snow_profile(ws_property, ic_property=None, reference_d={'ice': ['ice surface', 'down']}):
    """
    :param ws_property:
        openpyxl.worksheet
    :param ic_property:
    :param reference_d:
        top, or bottom
    """

    from pysic.core.profile import Profile
    from pysic.property import prop_associated

    logger = logging.getLogger(__name__)

    logger.error('%s: need to be implemented', ws_property)

    # find last column number with column header and/or subheaders
    if ws_property.max_row < MAX_ROW:
        max_row = ws_property.max_row
    else:
        max_row = MAX_ROW
    if ws_property.max_column < MAX_COL:
        max_col = ws_property.max_column
    else:
        max_col = MAX_COL
    min_row = 4


    # Dictionnary
    # parse property headers:
    # similar block entry as function read_generic_profile(), but adapted to read the 6 data blocks (salinity, temperature, nutrient, extra sample, SMP, SEW)

    super_header = {}
    super_header_unit_d = {}
    new_header_block = 0
    n_col_min = 1  # start column
    n_col = n_col_min
    empty_header = 0

    while new_header_block < 6:
        header_d = {}
        header_unit_d = {}

        n_row = 1  # header row

        empty_header = 0
        max_empty_header = 1

        while empty_header < 3 and n_col < max_col:
            # Read depth
            if isinstance(ws_property.cell(n_row, n_col).value, str):
                if 'depth' in ws_property.cell(n_row, n_col).value:
                    h_ = ws_property.cell(n_row, n_col).value
                    hs_ = ws_property.cell(n_row + 1, n_col).value
                    hu_ = ws_property.cell(n_row + 2, n_col).value
                    if h_ not in header_d:
                        header_d[h_] = {hs_: n_col}
                        header_unit_d[h_] = {hs_: hu_}
                    else:
                        header_d[h_].update({hs_: n_col})
                        header_unit_d[h_].update({hs_: hu_})
                elif ws_property.cell(n_row, n_col).value == 'comment':
                    h_ = ws_property.cell(n_row, n_col).value
                    hs_ = ws_property.cell(n_row + 1, n_col).value
                    hu_ = ws_property.cell(n_row + 2, n_col).value
                    header_d[h_] = {hs_: n_col}
                    header_unit_d[h_] = {hs_: hu_}

                # specific reader for temperature
                elif ws_property.cell(n_row, n_col).value == 'temperature':
                    h_ = ws_property.cell(n_row, n_col).value
                    hs_ = ws_property.cell(n_row + 1, n_col).value
                    hu_ = ws_property.cell(n_row + 2, n_col).value
                    if h_ not in header_d:
                        header_d[h_] = {hs_: n_col}
                        header_unit_d[h_] = {hs_: hu_}
                    else:
                        header_d[h_].update({hs_: n_col})
                        header_unit_d[h_] = {hs_: hu_}

                    hs_ = ws_property.cell(n_row + 1, n_col+1).value
                    hu_ = ws_property.cell(n_row + 2, n_col+1).value
                    header_d[h_].update({hs_: n_col+1})
                    header_unit_d[h_] = {hs_: hu_}

                # specific reader entry for eco_pool tab
                else:
                    prop_col = n_col
                    new_prop = False
                    qual_col = None
                    while not new_prop and empty_header < max_empty_header:
                        if isinstance(ws_property.cell(1, prop_col).value, str):
                            if 'ID' in header_d[h_] and 'quality' not in header_d[h_]:
                                if qual_col is not None:
                                    header_d[h_].update({'quality': qual_col})
                                else:
                                    _col = n_col
                                    while ws_property.cell(n_row + 1,
                                                           _col).value != 'quality' and _col < ws_property.max_column:
                                        _col += 1
                                    qual_col = _col
                                    if qual_col is not None:
                                        header_d[h_].update({'quality': qual_col})
                                    else:
                                        logger.error(
                                            'pysic.load.read_generic_profile: undefined quality column for property %s' % h_)
                            h_ = ws_property.cell(n_row, prop_col).value
                            hs_ = ws_property.cell(n_row + 1, prop_col).value
                            hu_ = ws_property.cell(n_row + 2, prop_col).value
                            header_d[h_] = {hs_: prop_col}
                            header_unit_d[h_] = {hs_: hu_}
                        elif ws_property.cell(2, prop_col).value == 'quality':
                            qual_col = prop_col
                            hs_ = ws_property.cell(n_row + 1, qual_col).value
                            header_d[h_].update({hs_: prop_col})
                        elif ws_property.cell(n_row + 1, prop_col).value != 'ID':
                            if ws_property.cell(n_row + 1, prop_col).value is not None:
                                hs_ = ws_property.cell(n_row + 1, prop_col).value
                                header_d[h_].update({hs_: prop_col})
                            else:
                                empty_header += 1
                        prop_col += 1

                        _ID_flag = (ws_property.cell(n_row + 1, prop_col).value == 'ID')
                        _comment_flag = (ws_property.cell(n_row, prop_col).value == 'comment')
                        _ish_flag =  (ws_property.cell(n_row, prop_col).value == 'ice section height')
                        if _ID_flag or _comment_flag or  _ish_flag:
                            if 'ID' in header_d[h_] and 'quality' not in header_d[h_]:
                                header_d[h_].update({'quality': qual_col})
                            new_prop = True
                            n_col = n_col + (prop_col - n_col -1)
                n_col += 1
            # search for subsamples ID
            else:
                prop_col = n_col
                new_prop = False
                h_ = None
                subheader_d = {}
                subheader_unit_d = {}
                while not new_prop and empty_header < max_empty_header and prop_col < max_col:
                    if isinstance(ws_property.cell(1, prop_col).value, str):
                        h_ = ws_property.cell(n_row, prop_col).value
                        hs_ = ws_property.cell(n_row + 1, prop_col).value
                        hu_ = ws_property.cell(n_row + 2, prop_col).value
                        header_d[h_] = {hs_: prop_col}
                        header_unit_d[h_] = {hs_: hu_}

                        if 'quality' not in header_d[h_] or 'quality' not in subheader_d:
                            _col = prop_col
                            while not ws_property.cell(n_row + 1, _col).value == 'ID' and _col < ws_property.max_column:
                                if ws_property.cell(n_row + 1, _col).value == 'quality':
                                    subheader_d['quality'] = _col
                                    subheader_unit_d['quality'] = ws_property.cell(n_row + 2, _col).value
                                    break
                                _col += 1

                        for key in subheader_d:
                            header_d[h_].update({key: subheader_d[key]})
                            header_unit_d[h_].update({key: subheader_unit_d[key]})

                    elif h_ is None:
                        hs_ = ws_property.cell(n_row + 1, prop_col).value
                        hu_ = ws_property.cell(n_row + 2, prop_col).value
                        subheader_d[hs_] = prop_col
                        subheader_unit_d[hs_] = hu_
                    else:
                        hs_ = ws_property.cell(n_row + 1, prop_col).value
                        hu_ = ws_property.cell(n_row + 2, prop_col).value
                        header_d[h_].update({hs_: prop_col})
                        header_unit_d[h_].update({hs_: hu_})
                        for key in subheader_d:
                            header_d[h_].update({key: subheader_d[key]})
                            header_unit_d[h_].update({key: subheader_unit_d[key]})
                    prop_col += 1

                    if ws_property.cell(n_row + 1, prop_col).value == 'ID' or ws_property.cell(n_row, prop_col).value == 'comment':
                        new_prop = True
                n_col = prop_col
                del subheader_d, subheader_unit_d, prop_col, new_prop

        super_header[new_header_block] = header_d
        super_header_unit_d[new_header_block] = header_unit_d
        new_header_block += 1

    del empty_header, n_col, n_row, hu_, h_

    # Read headers

    # Read Salinity block
    # read headers:
    n_row = 1
    n_col = 1
    n_col_min = n_col
    cell_flag = 2
    headers = []
    subheaders = []
    units = []
    while cell_flag >= 1:
        if isinstance(ws_property.cell(n_row, n_col).value, str):
            h_ = ws_property.cell(n_row, n_col).value
            headers.append(ws_property.cell(n_row, n_col).value)
            hs_ = ws_property.cell(n_row + 1, n_col).value
            subheaders.append(ws_property.cell(n_row + 1, n_col).value)
            units.append(ws_property.cell(n_row + 2, n_col).value)
            cell_flag = 2
            n_col += 1
        elif isinstance(ws_property.cell(n_row+1, n_col).value, str):
            headers.append(h_)
            subheaders.append(ws_property.cell(n_row+1, n_col).value)
            hs_ = ws_property.cell(n_row + 1, n_col).value
            units.append(ws_property.cell(n_row+2, n_col).value)
            n_col += 1
        else:
            cell_flag = 0
            cell_mark = n_col + 1
    n_col_max = n_col - n_col_min

    if ws_property.max_row < MAX_ROW:
        max_row = ws_property.max_row
    else:
        max_row = MAX_ROW
    min_row = 4

    # Check for step or continuous profiles:
    if 'depth center' in headers:
        loc1 = [ii for ii, h in enumerate(headers) if h == 'depth center'][0] + 1
        headers[loc1-1] = 'y_mid'
        y_mid = np.array(
            [ws_property.cell(row, loc1).value for row in range(min_row, max_row)]).astype(float)

        # discard trailing nan value from the end up
        # find nan value in y_low and y_sup
        y_nan_loc = np.where(np.isnan(y_mid))[0]
        # discard trailing nan value starting at the end
        if len(y_nan_loc) > 0 and len(y_mid) > 1 and y_nan_loc[-1] == len(y_mid)-1:
            y_nan_loc = [len(y_mid)-1] + [val for ii, val in enumerate(y_nan_loc[-2::-1]) if val == y_nan_loc[::-1][ii]-1]
            y_nan_loc = y_nan_loc[::-1]
            y_mid = [y for ii, y in enumerate(y_mid) if ii not in y_nan_loc]

    if 'depth 1' in headers and 'depth 2' in headers:
        step_flag = 1

        # find column with depth 1
        # TODO find a better way to find the location
        loc1 = [ii for ii, h in enumerate(headers) if h == 'depth 1'][0]+1
        headers[loc1 - 1] = 'y_low'
        loc2 = [ii for ii, h in enumerate(headers) if h == 'depth 2'][0]+1
        headers[loc2 - 1] = 'y_sup'
        # TODO: remove 'depth center'
        y_low = np.array([ws_property.cell(row, loc1).value for row in range(min_row, max_row + 1)]).astype(float)
        y_sup = np.array([ws_property.cell(row, loc2).value for row in range(min_row, max_row + 1)]).astype(float)

        # discard trailing nan value from the end up
        # find nan value in y_low and y_sup
        y_nan_loc = [ii for ii in np.where(np.isnan(y_sup))[0] if ii in np.where(np.isnan(y_low))[0]]

        # discard trailing nan value starting at the end
        if len(y_sup) > 0 and len(y_nan_loc) > 0 and y_nan_loc[-1] == len(y_sup)-1:
            y_nan_loc = [len(y_sup)-1] + [val for ii, val in enumerate(y_nan_loc[-2::-1]) if val == y_nan_loc[::-1][ii]-1]
            y_nan_loc = y_nan_loc[::-1]
            y_low = np.array([y for ii, y in enumerate(y_low) if ii not in y_nan_loc])
            y_sup = np.array([y for ii, y in enumerate(y_sup) if ii not in y_nan_loc])

            # TODO: replace missing y_low and y_sup with y_mid if possible

        if len(y_low) == 0:
            profile = Profile()
            y_mid = []
        elif 'y_mid' in headers:
            if np.isnan(y_mid).any() or len(y_mid) == 0:
                y_mid = (y_low + y_sup) / 2
                logger.info('(%s ) not all y_mid exits, calculating y_mid = (y_low+y_sup)/2'
                            % (ws_property.title))
            elif np.any(np.abs(((y_low + y_sup) / 2) - y_mid > 1e-12)):
                logger.error('(%s ) y_mid are not mid point between y_low and y_sup. \\'
                             'Replacing with y_mid = (y_low+y_sup)/2'
                             % (ws_property.title))
                y_mid = (y_low + y_sup) / 2
            else:
                logger.info('(%s ) y_low, y_mid and y_sup read with success'
                            % (ws_property.title))
        else:
            y_mid = (y_low + y_sup) / 2
    elif 'depth 1' in headers:
        loc1 = [ii for ii, h in enumerate(headers) if h == 'depth 1'][0]+1
        headers[loc1 - 1] = 'y_low'
        # TODO : fill y_sup
    elif 'depth 2' in headers:
        loc1 = [ii for ii, h in enumerate(headers) if h == 'depth 2'][0]+1
        headers[loc1 - 1] = 'y_sup'
        # TODO : fill y_low
    # Continuous profile
    else:
        y_low = np.nan * np.ones(y_mid.__len__())
        y_sup = np.nan * np.ones(y_mid.__len__())

    # Read data:
    # look up for first numeric or standard entry value
    # n_row_min = 4

    if len(y_mid) > 0:
        n_row_min = 1
        n_col_min = 1
        while not isinstance(ws_property.cell(n_row_min, n_col_min).value, (float, int)):
            n_row_min += 1
            if n_row_min > 1000:
                break
        n_row_max = n_row_min + len(y_mid) - 1

        # Drop column with depth:
        # _data = [[cell.value if isinstance(cell.value, (float, int)) else np.nan for cell in row]
        #                   for row in ws_property.iter_rows(n_row_min, n_row_max, n_col_min, n_col_max)]
        _data = [[cell.value for cell in row] for row in ws_property.iter_rows(n_row_min, n_row_max, n_col_min, n_col_max)]

        # TODO:  fill missing section with np.nan
        # if fill_missing:
        #     idx = np.where(np.abs(y_low[1:-1]-y_sup[0:-2]) > TOL)[0]
        #     for ii_idx in idx:
        #         empty = [y_sup[ii_idx], (y_sup[ii_idx]+y_low[ii_idx+1])/2, y_low[ii_idx+1]]
        #         empty += [np.nan] * (variable_headers.__len__()+1)
        #     data = np.vstack([data, empty])

        # concatenate header and subheader
        if ic_property is None:
            variable_prefix = ''
            phase = 'N/A'
        elif any(map(ic_property.__contains__, ['brine', 'sackhole'])):
            variable_prefix = 'brine_'
            phase = 'brine'
        elif any(map(ic_property.__contains__, ['seawater'])):
            variable_prefix = 'seawater_'
            phase = 'seawater'
        elif any(map(ic_property.__contains__, ['snow'])):
            variable_prefix = 'snow_'
            phase = 'snow'
        else:
            variable_prefix = ''
            phase = 'seaice'
        subheaders = [sh if sh is not None else '' for sh in subheaders]
        profile_headers = [variable_prefix + h + '_' + subheaders[ii] if (len(subheaders[ii]) > 1 and h not in ['y_low', 'y_sup', 'y_mid']) else h
                           for ii, h in enumerate(headers)]
        # TODO: double header for dataframe with header and subheader

        profile = pd.DataFrame(_data, columns=profile_headers)
        if 'y_mid' not in profile.keys():
            profile['y_mid'] = y_mid

        # drop empty variable header
        if None in profile.columns:
            profile = profile.drop(labels=[None], axis=1)

        # sample ID columns is string:
        string_header = ['comment'] + [h for h in profile.columns if 'sample ID' in h or 'ID' in h]

        # convert string to float:
        float_header = [h for h in profile.columns if h not in string_header]
        profile[float_header] = profile[float_header].apply(pd.to_numeric, errors='coerce')

        # drop property with all nan value
        profile = profile.dropna(axis=1, how='all')

        # remove empty line if all element of depth are nan:
        subset = [col for col in ['y_low', 'y_sup', 'y_mid'] if col in profile.columns]
        profile = profile.dropna(axis=0, subset=subset, how='all')

        # singularize comments
        if 'comments' in profile.columns:
            profile.rename(columns={'comments': "comment"}, inplace=True)
        # add comment column if it does not exist
        if 'comment' not in profile.columns:
            profile['comment'] = None
        else:
            profile['comment'] = profile['comment'].astype(str).replace({'nan': None})

        # get all property variable (e.g. salinity, temperature, ...)
        property = [var for var in profile.columns if var not in ['comment', 'y_low', 'y_sup', 'y_mid']]
        property = [prop.split('_')[0] for prop in property]
        property = list(set(property))

        # remove subvariable (e.g. conductivity temperature measurement for conductivity
        property = [prop for prop in property if prop not in inverse_dict(subvariable_dict)]

        # set variable to string of property
        profile['variable'] = [', '.join(property)] * len(profile.index)

        # set vertical references
        # TODO: improve vertical reference
        if reference_d['snow'][0] == 'ice surface':
            v_ref = 'bottom'
        elif reference_d['snow'][0] == 'snow interface':
            v_ref = 'bottom'
        else:
            logger.error(ws_property.title + ' - Vertical references not set or not yet handled')
        profile['v_ref'] = [v_ref] * len(profile.index)

        # set columns type
        col_string = ['comment', 'v_ref', 'name', 'profile', 'variable']
        col_date = ['date']
        col_float = [h for h in profile.columns if h not in col_string and h not in col_date and 'ID' not in h]
        col_string = col_string + [h for h in profile.columns if 'ID' in h]
        profile[col_float] = profile[col_float].apply(pd.to_numeric, errors='coerce')
        c_string = [h for h in col_string if h in profile.columns]
        profile[c_string] = profile[c_string].astype(str).replace({'nan': None})

        profile = Profile(profile)
        # remove variable not in variables
        if ic_property is not None:
            for property in profile.properties():
                if property not in ic_property:
                    profile.delete_property(property)

    if ic_property is not None and 'salinity' not in ic_property:
        profile_S = Profile()
    else:
        profile_S = profile

    cell_mark = 26

    del profile
    # Read Temperature block
    # read headers:
    n_row = 1
    n_col = cell_mark
    n_col_min = n_col
    cell_flag = 2
    headers = []
    subheaders = []
    units = []
    while cell_flag >= 1:
        if isinstance(ws_property.cell(n_row, n_col).value, str):
            h_ = ws_property.cell(n_row, n_col).value
            headers.append(ws_property.cell(n_row, n_col).value)
            hs_ = ws_property.cell(n_row + 1, n_col).value
            subheaders.append(ws_property.cell(n_row + 1, n_col).value)
            units.append(ws_property.cell(n_row + 2, n_col).value)
            cell_flag = 2
            n_col += 1
        elif isinstance(ws_property.cell(n_row + 1, n_col).value, str):
            headers.append(h_)
            hs_ = ws_property.cell(n_row + 1, n_col).value
            subheaders.append(hs_)
            units.append(ws_property.cell(n_row + 2, n_col).value)
            n_col += 1
        else:
            cell_flag = 0
            cell_mark = n_col + 1
    n_col_max = n_col

    if ws_property.max_row < MAX_ROW:
        max_row = ws_property.max_row
    else:
        max_row = MAX_ROW
    min_row = 4

    # Check for step or continuous profiles:
    #loc1 = [ii for ii, h in enumerate(headers) if h == 'depth'][0] + 1
    headers[loc1 - 1] = 'depth'
    subheaders[loc1 - 1] = 'y_mid'
    headers = ['y_mid', 'temperature', 'temperature']
    subheaders = ['', 'value', 'quality']
    y_mid = np.array(
        [ws_property.cell(row, n_col_min).value for row in range(min_row, max_row)]).astype(float)

    # discard trailing nan value from the end up
    # find nan value in y_low and y_sup
    y_nan_loc = np.where(np.isnan(y_mid))[0]
    # discard trailing nan value starting at the end
    if len(y_nan_loc) > 0 and len(y_mid) > 1 and y_nan_loc[-1] == len(y_mid) - 1:
        y_nan_loc = [len(y_mid) - 1] + [val for ii, val in enumerate(y_nan_loc[-2::-1]) if
                                        val == y_nan_loc[::-1][ii] - 1]
        y_nan_loc = y_nan_loc[::-1]
        y_mid = [y for ii, y in enumerate(y_mid) if ii not in y_nan_loc]

    y_low = np.nan * np.ones(y_mid.__len__())
    y_sup = np.nan * np.ones(y_mid.__len__())

    # Read data:
    # look up for first numeric or standard entry value
    # n_row_min = 4

    if len(y_mid) > 0:
        n_row_min = 1
        while not isinstance(ws_property.cell(n_row_min, n_col_min).value, (float, int)):
            n_row_min += 1
            if n_row_min > 1000:
                break
        n_row_max = n_row_min + len(y_mid) - 1

        # Drop column with depth:
        # _data = [[cell.value if isinstance(cell.value, (float, int)) else np.nan for cell in row]
        #                   for row in ws_property.iter_rows(n_row_min, n_row_max, n_col_min, n_col_max)]
        _data = [[cell.value for cell in row] for row in
                 ws_property.iter_rows(n_row_min, n_row_max, n_col_min, n_col_max-1)]

        # TODO:  fill missing section with np.nan
        # if fill_missing:
        #     idx = np.where(np.abs(y_low[1:-1]-y_sup[0:-2]) > TOL)[0]
        #     for ii_idx in idx:
        #         empty = [y_sup[ii_idx], (y_sup[ii_idx]+y_low[ii_idx+1])/2, y_low[ii_idx+1]]
        #         empty += [np.nan] * (variable_headers.__len__()+1)
        #     data = np.vstack([data, empty])

        # concatenate header and subheader
        if ic_property is None:
            variable_prefix = ''
            phase = 'N/A'
        elif any(map(ic_property.__contains__, ['brine', 'sackhole'])):
            variable_prefix = 'brine_'
            phase = 'brine'
        elif any(map(ic_property.__contains__, ['seawater'])):
            variable_prefix = 'seawater_'
            phase = 'seawater'
        elif any(map(ic_property.__contains__, ['snow'])):
            variable_prefix = 'snow_'
            phase = 'snow'
        else:
            variable_prefix = ''
            phase = 'seaice'
        subheaders = [sh if sh is not None else '' for sh in subheaders]
        profile_headers = [variable_prefix + h + '_' + subheaders[ii] if (
                    len(subheaders[ii]) > 1 and h not in ['y_low', 'y_sup', 'y_mid']) else h
                           for ii, h in enumerate(headers)]
        # TODO: double header for dataframe with header and subheader

        profile = pd.DataFrame(_data, columns=profile_headers)
        if 'y_mid' not in profile.keys():
            profile['y_mid'] = y_mid

        # drop empty variable header
        if None in profile.columns:
            profile = profile.drop(labels=[None], axis=1)

        # sample ID columns is string:
        string_header = ['comment'] + [h for h in profile.columns if 'sample ID' in h or 'ID' in h]

        # convert string to float:
        float_header = [h for h in profile.columns if h not in string_header]
        profile[float_header] = profile[float_header].apply(pd.to_numeric, errors='coerce')

        # drop property with all nan value
        profile = profile.dropna(axis=1, how='all')

        # remove empty line if all element of depth are nan:
        subset = [col for col in ['y_low', 'y_sup', 'y_mid'] if col in profile.columns]
        profile = profile.dropna(axis=0, subset=subset, how='all')

        # singularize comments
        if 'comments' in profile.columns:
            profile.rename(columns={'comments': "comment"}, inplace=True)
        # add comment column if it does not exist
        if 'comment' not in profile.columns:
            profile['comment'] = None
        else:
            profile['comment'] = profile['comment'].astype(str).replace({'nan': None})

        # get all property variable (e.g. salinity, temperature, ...)
        property = [var for var in profile.columns if var not in ['comment', 'y_low', 'y_sup', 'y_mid']]
        property = [prop.split('_')[0] for prop in property]
        property = list(set(property))

        # remove subvariable (e.g. conductivity temperature measurement for conductivity
        property = [prop for prop in property if prop not in inverse_dict(subvariable_dict)]

        # set variable to string of property
        profile['variable'] = [', '.join(property)] * len(profile.index)

        # set vertical references
        # TODO: improve vertical reference
        if reference_d['snow'][0] == 'ice surface':
            v_ref = 'bottom'
        elif reference_d['snow'][0] == 'snow interface':
            v_ref = 'bottom'
        else:
            logger.error(ws_property.title + ' - Vertical references not set or not yet handled')
        profile['v_ref'] = [v_ref] * len(profile.index)

        # set columns type
        col_string = ['comment', 'v_ref', 'name', 'profile', 'variable']
        col_date = ['date']
        col_float = [h for h in profile.columns if h not in col_string and h not in col_date and 'ID' not in h]
        col_string = col_string + [h for h in profile.columns if 'ID' in h]
        profile[col_float] = profile[col_float].apply(pd.to_numeric, errors='coerce')
        c_string = [h for h in col_string if h in profile.columns]
        profile[c_string] = profile[c_string].astype(str).replace({'nan': None})

        profile = Profile(profile)
        # remove variable not in variables
        if ic_property is not None:
            for _property in profile.properties():
                if _property not in ic_property:
                    profile.delete_property(_property)

        if ic_property is not None and 'temperature' not in ic_property:
            profile = profile_S
        else:
            profile = profile.append(profile_S)
        profile.reset_index(drop=True, inplace=True)
    else:
        profile = Profile()
    return profile


def read_metadata_variable_as_str(ws, variable, col_header_idx=1, col_variable_idx=3, row_offset=0):
    row_idx = find_str_in_col(ws, variable, col_header_idx)[0] + row_offset
    if ws.cell(row_idx, col_variable_idx).value is not None:
        return ws.cell(row_idx, col_variable_idx).value
    else:
        return 'N/A'
def read_metadata_variable_as_float(ws, variable, col_header_idx=1, col_variable_idx=3, row_offset=0):
    row_idx = find_str_in_col(ws, variable, col_header_idx)[0] + row_offset
    variable_value = ws.cell(row_idx, col_variable_idx).value
    if isfloat(variable_value):
        return float(variable_value)
    else:
        return np.nan
